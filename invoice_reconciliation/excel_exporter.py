from __future__ import annotations

from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import ReconciliationResult, ReconciliationStatus


class ExcelExporter(object):
    """
    Export ket qua doi chieu ra file Excel.
    """

    HEADERS = (
        "Receipt No",
        "Item",
        "Description",
        "Receipt Date",
        "Quantity",
        "Price",
        "Amount",
        "P/O No",
        "Note",
    )

    PASS_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")
    MISMATCH_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
    MISSING_FILL = PatternFill(fill_type="solid", fgColor="F8CBAD")
    HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9E1F2")

    def export(self, result, output_path):
        """
        :param ReconciliationResult result:
        :param str output_path:
        :return: output_path
        :rtype: str
        """
        workbook = Workbook()
        reconciliation_sheet = workbook.active
        reconciliation_sheet.title = "Reconciliation"
        self._write_reconciliation_sheet(reconciliation_sheet, result)

        summary_sheet = workbook.create_sheet("Summary")
        self._write_summary_sheet(summary_sheet, result)

        workbook.save(output_path)
        return output_path

    def _write_reconciliation_sheet(self, sheet, result):
        sheet.append(list(self.HEADERS))
        self._style_header(sheet, len(self.HEADERS))

        missing_rows = []
        missing_total = Decimal("0")

        for row in result.rows:
            sheet.append(
                [
                    row.receipt_no,
                    row.item_code,
                    row.description,
                    row.receipt_date,
                    row.receipt_qty,
                    row.receipt_unit_price,
                    row.receipt_amount,
                    row.po_no,
                    row.note,
                ]
            )

            fill = self._status_fill(row.final_status)
            excel_row_index = sheet.max_row
            for col in range(1, len(self.HEADERS) + 1):
                cell = sheet.cell(row=excel_row_index, column=col)
                cell.fill = fill
                if col in (5, 6, 7) and cell.value is not None:
                    cell.number_format = "#,##0.00"

            if row.final_status == ReconciliationStatus.MISSING_IN_INVOICE:
                row_missing_amount = row.missing_amount if row.missing_amount is not None else row.receipt_amount
                if row_missing_amount is not None:
                    missing_total += Decimal(row_missing_amount)
                missing_rows.append(row)

        self._append_missing_summary(sheet, missing_rows, missing_total)

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = "A1:I%s" % sheet.max_row
        self._auto_fit_columns(sheet, len(self.HEADERS))

    def _write_summary_sheet(self, sheet, result):
        summary = result.summary
        sheet.append(["Metric", "Value"])
        self._style_header(sheet, 2)
        sheet.append(["Total Invoice Amount", summary.total_invoice_amount])
        sheet.append(["Total Receipt Amount", summary.total_receipt_amount])
        sheet.append(["Matched Count", summary.matched_count])
        sheet.append(["Mismatch Count", summary.mismatch_count])
        sheet.append(["Missing Count", summary.missing_count])
        sheet["B2"].number_format = "#,##0.00"
        sheet["B3"].number_format = "#,##0.00"
        self._auto_fit_columns(sheet, 2)

    def _status_fill(self, status):
        if status == ReconciliationStatus.PASS:
            return self.PASS_FILL
        if status in (ReconciliationStatus.MISSING_IN_INVOICE, ReconciliationStatus.MISSING_IN_RECEIPT):
            return self.MISSING_FILL
        return self.MISMATCH_FILL

    def _append_missing_summary(self, sheet, missing_rows, missing_total):
        start_row = sheet.max_row + 2
        sheet.cell(row=start_row, column=1, value="Tong ket phan chua xuat hoa don")
        sheet.cell(row=start_row, column=1).font = Font(bold=True)

        headers = ("Item", "Description", "Quantity", "Price", "Missing Amount", "P/O No")
        for index, title in enumerate(headers, start=1):
            cell = sheet.cell(row=start_row + 1, column=index, value=title)
            cell.font = Font(bold=True)
            cell.fill = self.HEADER_FILL

        write_row = start_row + 2
        for row in missing_rows:
            sheet.cell(row=write_row, column=1, value=row.item_code)
            sheet.cell(row=write_row, column=2, value=row.description)
            sheet.cell(row=write_row, column=3, value=row.receipt_qty)
            sheet.cell(row=write_row, column=4, value=row.receipt_unit_price)
            amount = row.missing_amount if row.missing_amount is not None else row.receipt_amount
            sheet.cell(row=write_row, column=5, value=amount)
            sheet.cell(row=write_row, column=6, value=row.po_no)
            for col in (3, 4, 5):
                sheet.cell(row=write_row, column=col).number_format = "#,##0.00"
            write_row += 1

        total_label_row = write_row
        sheet.cell(row=total_label_row, column=4, value="Total Missing")
        sheet.cell(row=total_label_row, column=4).font = Font(bold=True)
        sheet.cell(row=total_label_row, column=5, value=missing_total)
        sheet.cell(row=total_label_row, column=5).font = Font(bold=True)
        sheet.cell(row=total_label_row, column=5).number_format = "#,##0.00"

    def _style_header(self, sheet, total_columns):
        for col in range(1, total_columns + 1):
            cell = sheet.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.fill = self.HEADER_FILL

    def _auto_fit_columns(self, sheet, total_columns):
        for col in range(1, total_columns + 1):
            letter = get_column_letter(col)
            max_len = 0
            for row in range(1, sheet.max_row + 1):
                value = sheet.cell(row=row, column=col).value
                text = "" if value is None else str(value)
                if len(text) > max_len:
                    max_len = len(text)
            sheet.column_dimensions[letter].width = min(max_len + 2, 50)
